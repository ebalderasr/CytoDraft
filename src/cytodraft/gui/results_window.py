from __future__ import annotations

import csv
import re
from pathlib import Path

import numpy as np
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QButtonGroup,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QSizePolicy,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from cytodraft.core.export import export_masked_events_to_csv, export_masked_events_to_fcs
from cytodraft.core.statistics import STATISTIC_DEFINITIONS
from cytodraft.models.workspace import COMPENSATION_GROUP_NAME, WorkspaceState
from cytodraft.services.statistics_service import StatisticsService

_DEFAULT_METRIC_KEYS = {"event_count", "percent_parent", "percent_total", "mean", "median"}

_SAMPLE_BG = QColor("#f7f7ff")
_GROUP_BG = QColor("#eefbf3")
_STAT_BG = QColor("#f5f3ff")
_MISSING_BG = QColor("#f3f4f6")
_MISSING_FG = QColor("#9ca3af")


def _safe_filename(name: str) -> str:
    """Remove filesystem-unsafe characters from a name."""
    return re.sub(r"[^\w\-_.]", "_", name).strip("_") or "export"


def _write_stats_xlsx(path: str, headers: list[str], rows: list[list]) -> None:
    """Write the statistics table to an .xlsx file with column color formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Column fill colours (match the Qt table)
    _FILL_SAMPLE = PatternFill("solid", fgColor="F7F7FF")
    _FILL_GROUP = PatternFill("solid", fgColor="EEFBF3")
    _FILL_STAT = PatternFill("solid", fgColor="F5F3FF")
    _FILL_HEADER = PatternFill("solid", fgColor="EDE9FE")
    _FONT_HEADER = Font(bold=True)
    _ALIGN_RIGHT = Alignment(horizontal="right")
    _ALIGN_CENTER = Alignment(horizontal="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Statistics"

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER

    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.fill = _FILL_SAMPLE
            elif col_idx == 2:
                cell.fill = _FILL_GROUP
            else:
                cell.fill = _FILL_STAT
                cell.alignment = _ALIGN_RIGHT

    # Auto-fit column widths (cap at 40)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, len(rows) + 2)),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(path)


class ResultsWindow(QDialog):
    """Statistics preview, batch events export, and CSV/FCS export."""

    def __init__(
        self,
        workspace: WorkspaceState,
        statistics_service: StatisticsService,
        parent=None,
    ) -> None:
        super().__init__(parent)
        self.workspace = workspace
        self.statistics_service = statistics_service

        self.setWindowTitle("Results")
        self.setMinimumSize(860, 540)
        self.resize(1200, 720)
        self.setWindowFlags(
            Qt.Window
            | Qt.WindowCloseButtonHint
            | Qt.WindowMinimizeButtonHint
            | Qt.WindowMaximizeButtonHint
        )

        self._build_ui()
        self._connect_signals()
        self.refresh()

    # ------------------------------------------------------------------
    # Shared helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _vline() -> QFrame:
        sep = QFrame()
        sep.setFrameShape(QFrame.VLine)
        sep.setFrameShadow(QFrame.Plain)
        sep.setFixedWidth(1)
        sep.setStyleSheet("background: #d4dde7; border: none;")
        sep.setContentsMargins(0, 4, 0, 4)
        return sep

    @staticmethod
    def _label(text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet("color: #536274; font-weight: 600; font-size: 12px;")
        return lbl

    @staticmethod
    def _hint(text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setWordWrap(True)
        lbl.setStyleSheet("color: #6b7280; font-size: 11px;")
        return lbl

    def _non_compensation_samples(
        self, group_name: str | None
    ) -> list[tuple[int, object]]:
        return [
            (idx, ws)
            for idx, ws in self.workspace.samples_in_group(group_name)
            if ws.group_name != COMPENSATION_GROUP_NAME
        ]

    def _collect_groups_populations_channels(
        self, group_name: str | None
    ) -> tuple[list[str], list[str]]:
        """Return (populations, channels) for non-compensation samples in group."""
        populations: list[str] = ["All events"]
        seen_pops: set[str] = {"All events"}
        channels: list[str] = []
        seen_channels: set[str] = set()

        for _, ws in self._non_compensation_samples(group_name):
            for gate in ws.gates:
                if gate.name not in seen_pops:
                    populations.append(gate.name)
                    seen_pops.add(gate.name)
            for ch in ws.sample.channels:
                if ch.display_name not in seen_channels:
                    channels.append(ch.display_name)
                    seen_channels.add(ch.display_name)

        return populations, channels

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _build_ui(self) -> None:
        self._tabs = QTabWidget()
        self._tabs.addTab(self._build_statistics_tab(), "Statistics")
        self._tabs.addTab(self._build_events_tab(), "Events")

        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(0)
        layout.addWidget(self._tabs)
        self.setLayout(layout)

    # ── Statistics tab ────────────────────────────────────────────────

    def _build_statistics_tab(self) -> QWidget:
        tab = QWidget()

        # Controls row
        controls = QHBoxLayout()
        controls.setSpacing(8)
        controls.setContentsMargins(0, 0, 0, 0)

        self._stats_group_combo = QComboBox()
        self._stats_group_combo.setMinimumWidth(140)

        self._stats_pop_combo = QComboBox()
        self._stats_pop_combo.setMinimumWidth(150)

        self._stats_channel_combo = QComboBox()
        self._stats_channel_combo.setMinimumWidth(150)

        self._stats_refresh_btn = QPushButton("↻ Refresh")
        self._stats_refresh_btn.setProperty("variant", "subtle")

        self._stats_export_btn = QPushButton("Export CSV")
        self._stats_export_btn.setProperty("variant", "primary")

        self._stats_export_xlsx_btn = QPushButton("Export Excel")
        self._stats_export_xlsx_btn.setProperty("variant", "subtle")
        self._stats_export_xlsx_btn.setToolTip("Export statistics table as .xlsx with formatting")

        controls.addWidget(self._label("Group:"))
        controls.addWidget(self._stats_group_combo)
        controls.addWidget(self._vline())
        controls.addWidget(self._label("Population:"))
        controls.addWidget(self._stats_pop_combo)
        controls.addWidget(self._vline())
        controls.addWidget(self._label("Channel:"))
        controls.addWidget(self._stats_channel_combo)
        controls.addStretch(1)
        controls.addWidget(self._stats_refresh_btn)
        controls.addWidget(self._stats_export_xlsx_btn)
        controls.addWidget(self._stats_export_btn)

        # Metrics row
        metrics_row = QHBoxLayout()
        metrics_row.setSpacing(6)
        metrics_row.setContentsMargins(0, 0, 0, 0)

        metrics_label = self._label("Metrics:")
        metrics_label.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)

        self._metric_list = QListWidget()
        self._metric_list.setSelectionMode(QListWidget.SelectionMode.NoSelection)
        self._metric_list.setFlow(QListWidget.Flow.LeftToRight)
        self._metric_list.setWrapping(True)
        self._metric_list.setFixedHeight(52)
        self._metric_list.setFrameShape(QFrame.NoFrame)
        self._metric_list.setStyleSheet(
            "QListWidget { background: transparent; } "
            "QListWidget::item { padding: 2px 8px; border-radius: 6px; } "
            "QListWidget::item:hover { background: #e0e7ff; }"
        )

        for key, label in STATISTIC_DEFINITIONS:
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, key)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if key in _DEFAULT_METRIC_KEYS else Qt.Unchecked)
            self._metric_list.addItem(item)

        metrics_row.addWidget(metrics_label)
        metrics_row.addWidget(self._metric_list, stretch=1)

        # Hint
        hint_row = QHBoxLayout()
        hint_row.setContentsMargins(0, 0, 0, 0)
        hint_row.addWidget(
            self._hint(
                "Las métricas de canal (Mean, Median, etc.) usan el canal seleccionado. "
                "Event count y porcentajes son independientes del canal."
            ),
            stretch=1,
        )
        hint_row.addWidget(
            self._hint("Gris: muestra  |  Verde: grupo  |  Violeta: estadísticas")
        )

        # Stats table
        self._stats_table = QTableWidget()
        self._stats_table.setAlternatingRowColors(True)
        self._stats_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._stats_table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self._stats_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._stats_table.verticalHeader().setVisible(False)
        self._stats_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive
        )
        self._stats_table.horizontalHeader().setStretchLastSection(False)
        self._stats_table.setWordWrap(False)

        layout = QVBoxLayout()
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(6)
        layout.addLayout(controls)
        layout.addLayout(metrics_row)
        layout.addLayout(hint_row)
        layout.addWidget(self._stats_table, stretch=1)
        tab.setLayout(layout)
        return tab

    # ── Events tab ────────────────────────────────────────────────────

    def _build_events_tab(self) -> QWidget:
        tab = QWidget()

        # ── Left panel: controls + sample list ────────────────────────
        left = QWidget()

        # Controls
        controls = QHBoxLayout()
        controls.setSpacing(8)
        controls.setContentsMargins(0, 0, 0, 0)

        self._ev_group_combo = QComboBox()
        self._ev_group_combo.setMinimumWidth(140)

        self._ev_gate_combo = QComboBox()
        self._ev_gate_combo.setMinimumWidth(150)

        controls.addWidget(self._label("Group:"))
        controls.addWidget(self._ev_group_combo)
        controls.addWidget(self._vline())
        controls.addWidget(self._label("Gate:"))
        controls.addWidget(self._ev_gate_combo)
        controls.addStretch(1)

        # Selection buttons
        sel_row = QHBoxLayout()
        sel_row.setSpacing(6)
        sel_row.setContentsMargins(0, 0, 0, 0)

        self._ev_select_all_btn = QPushButton("Select all")
        self._ev_select_all_btn.setProperty("variant", "subtle")
        self._ev_deselect_all_btn = QPushButton("Deselect all")
        self._ev_deselect_all_btn.setProperty("variant", "subtle")
        self._ev_select_ready_btn = QPushButton("Select with gate")
        self._ev_select_ready_btn.setProperty("variant", "subtle")
        self._ev_select_ready_btn.setToolTip(
            "Select only samples that have the chosen gate"
        )

        sel_row.addWidget(self._ev_select_all_btn)
        sel_row.addWidget(self._ev_deselect_all_btn)
        sel_row.addWidget(self._ev_select_ready_btn)
        sel_row.addStretch(1)

        # Sample list
        self._ev_sample_table = QTableWidget()
        self._ev_sample_table.setColumnCount(3)
        self._ev_sample_table.setHorizontalHeaderLabels(["Sample", "Group", "Events"])
        self._ev_sample_table.setAlternatingRowColors(True)
        self._ev_sample_table.setSelectionMode(
            QAbstractItemView.SelectionMode.NoSelection
        )
        self._ev_sample_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._ev_sample_table.verticalHeader().setVisible(False)
        self._ev_sample_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive
        )
        self._ev_sample_table.horizontalHeader().setStretchLastSection(True)
        self._ev_sample_table.setWordWrap(False)

        left_layout = QVBoxLayout()
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(6)
        left_layout.addLayout(controls)
        left_layout.addLayout(sel_row)
        left_layout.addWidget(self._ev_sample_table, stretch=1)
        left.setLayout(left_layout)

        # ── Right panel: summary + export options ─────────────────────
        right = QWidget()
        right.setMaximumWidth(280)

        self._ev_summary_label = QLabel("—")
        self._ev_summary_label.setWordWrap(True)
        self._ev_summary_label.setStyleSheet(
            "color: #18212f; font-size: 13px; padding: 8px;"
            "background: #f7f9fc; border: 1px solid #d7e0ea; border-radius: 8px;"
        )

        # Export mode radio buttons
        mode_label = self._label("Modo de exportación:")
        self._ev_radio_per_sample = QRadioButton("Un archivo por muestra")
        self._ev_radio_combined = QRadioButton("Un archivo combinado (CSV)")
        self._ev_radio_per_sample.setChecked(True)

        self._ev_radio_group = QButtonGroup(self)
        self._ev_radio_group.addButton(self._ev_radio_per_sample)
        self._ev_radio_group.addButton(self._ev_radio_combined)

        # Export buttons
        self._ev_export_csv_btn = QPushButton("Export CSV")
        self._ev_export_csv_btn.setProperty("variant", "primary")
        self._ev_export_fcs_btn = QPushButton("Export FCS")
        self._ev_export_fcs_btn.setProperty("variant", "subtle")

        export_row = QHBoxLayout()
        export_row.setSpacing(6)
        export_row.setContentsMargins(0, 0, 0, 0)
        export_row.addWidget(self._ev_export_csv_btn)
        export_row.addWidget(self._ev_export_fcs_btn)

        fcs_hint = self._hint(
            "FCS siempre exporta un archivo por muestra (formato estándar)."
        )

        right_layout = QVBoxLayout()
        right_layout.setContentsMargins(8, 0, 0, 0)
        right_layout.setSpacing(8)
        right_layout.addWidget(self._label("Resumen:"))
        right_layout.addWidget(self._ev_summary_label)
        right_layout.addSpacing(8)
        right_layout.addWidget(mode_label)
        right_layout.addWidget(self._ev_radio_per_sample)
        right_layout.addWidget(self._ev_radio_combined)
        right_layout.addSpacing(12)
        right_layout.addLayout(export_row)
        right_layout.addWidget(fcs_hint)
        right_layout.addStretch(1)
        right.setLayout(right_layout)

        # ── Splitter ──────────────────────────────────────────────────
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        splitter.addWidget(left)
        splitter.addWidget(right)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([860, 280])

        layout = QVBoxLayout()
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(6)
        layout.addWidget(
            self._hint(
                "Selecciona un gate y las muestras a exportar. "
                "Las muestras sin el gate aparecen en gris y no se exportan."
            )
        )
        layout.addWidget(splitter, stretch=1)
        tab.setLayout(layout)
        return tab

    # ------------------------------------------------------------------
    # Signal connections
    # ------------------------------------------------------------------

    def _connect_signals(self) -> None:
        # Statistics tab
        self._stats_group_combo.currentIndexChanged.connect(self._on_stats_group_changed)
        self._stats_pop_combo.currentIndexChanged.connect(self._rebuild_stats_table)
        self._stats_channel_combo.currentIndexChanged.connect(self._rebuild_stats_table)
        self._metric_list.itemChanged.connect(self._rebuild_stats_table)
        self._stats_refresh_btn.clicked.connect(self.refresh)
        self._stats_export_btn.clicked.connect(self._on_export_stats_csv)
        self._stats_export_xlsx_btn.clicked.connect(self._on_export_stats_xlsx)

        # Events tab
        self._ev_group_combo.currentIndexChanged.connect(self._on_ev_group_changed)
        self._ev_gate_combo.currentIndexChanged.connect(self._rebuild_ev_sample_list)
        self._ev_select_all_btn.clicked.connect(self._on_ev_select_all)
        self._ev_deselect_all_btn.clicked.connect(self._on_ev_deselect_all)
        self._ev_select_ready_btn.clicked.connect(self._on_ev_select_ready)
        self._ev_sample_table.itemChanged.connect(self._update_ev_summary)
        self._ev_export_csv_btn.clicked.connect(self._on_export_events_csv)
        self._ev_export_fcs_btn.clicked.connect(self._on_export_events_fcs)

    # ------------------------------------------------------------------
    # Refresh
    # ------------------------------------------------------------------

    def refresh(self) -> None:
        """Rebuild both tabs from current workspace state."""
        self._refresh_stats_tab()
        self._refresh_events_tab()

    def _refresh_stats_tab(self) -> None:
        current_group = self._stats_group_combo.currentData()
        self._stats_group_combo.blockSignals(True)
        self._stats_group_combo.clear()
        self._stats_group_combo.addItem("All groups", None)
        for name in sorted(self.workspace.groups):
            if name != COMPENSATION_GROUP_NAME:
                self._stats_group_combo.addItem(name, name)

        restored = any(
            self._stats_group_combo.itemData(i) == current_group
            and (self._stats_group_combo.setCurrentIndex(i) or True)
            for i in range(self._stats_group_combo.count())
        )
        if not restored:
            self._stats_group_combo.setCurrentIndex(0)
        self._stats_group_combo.blockSignals(False)

        self._rebuild_stats_population_channel()
        self._rebuild_stats_table()

    def _refresh_events_tab(self) -> None:
        current_group = self._ev_group_combo.currentData()
        self._ev_group_combo.blockSignals(True)
        self._ev_group_combo.clear()
        self._ev_group_combo.addItem("All groups", None)
        for name in sorted(self.workspace.groups):
            if name != COMPENSATION_GROUP_NAME:
                self._ev_group_combo.addItem(name, name)

        restored = any(
            self._ev_group_combo.itemData(i) == current_group
            and (self._ev_group_combo.setCurrentIndex(i) or True)
            for i in range(self._ev_group_combo.count())
        )
        if not restored:
            self._ev_group_combo.setCurrentIndex(0)
        self._ev_group_combo.blockSignals(False)

        self._rebuild_ev_gates()
        self._rebuild_ev_sample_list()

    # ------------------------------------------------------------------
    # Statistics tab logic
    # ------------------------------------------------------------------

    def _on_stats_group_changed(self) -> None:
        self._rebuild_stats_population_channel()
        self._rebuild_stats_table()

    def _rebuild_stats_population_channel(self) -> None:
        group_name = self._stats_group_combo.currentData()
        current_pop = self._stats_pop_combo.currentText()
        current_channel = self._stats_channel_combo.currentText()

        populations, channels = self._collect_groups_populations_channels(group_name)

        self._stats_pop_combo.blockSignals(True)
        self._stats_pop_combo.clear()
        self._stats_pop_combo.addItems(populations)
        idx = self._stats_pop_combo.findText(current_pop)
        if idx >= 0:
            self._stats_pop_combo.setCurrentIndex(idx)
        self._stats_pop_combo.blockSignals(False)

        self._stats_channel_combo.blockSignals(True)
        self._stats_channel_combo.clear()
        self._stats_channel_combo.addItems(channels)
        idx = self._stats_channel_combo.findText(current_channel)
        if idx >= 0:
            self._stats_channel_combo.setCurrentIndex(idx)
        self._stats_channel_combo.blockSignals(False)

    def _selected_metrics(self) -> list[tuple[str, str]]:
        result = []
        for i in range(self._metric_list.count()):
            item = self._metric_list.item(i)
            if item.checkState() == Qt.Checked:
                result.append((item.data(Qt.UserRole), item.text()))
        return result

    def _rebuild_stats_table(self) -> None:
        group_name = self._stats_group_combo.currentData()
        population_name = self._stats_pop_combo.currentText()
        channel_name = self._stats_channel_combo.currentText()
        metrics = self._selected_metrics()

        self._stats_table.blockSignals(True)

        if not population_name or not metrics:
            self._stats_table.setRowCount(0)
            self._stats_table.setColumnCount(0)
            self._stats_table.blockSignals(False)
            return

        samples = self._non_compensation_samples(group_name)
        headers = ["Sample", "Group"] + [label for _, label in metrics]
        self._stats_table.setColumnCount(len(headers))
        self._stats_table.setHorizontalHeaderLabels(headers)
        self._stats_table.setRowCount(len(samples))

        for row, (_, ws) in enumerate(samples):
            sample_item = QTableWidgetItem(ws.sample_name)
            sample_item.setBackground(_SAMPLE_BG)
            sample_item.setFlags(sample_item.flags() & ~Qt.ItemIsEditable)
            self._stats_table.setItem(row, 0, sample_item)

            group_item = QTableWidgetItem(ws.group_name)
            group_item.setBackground(_GROUP_BG)
            group_item.setFlags(group_item.flags() & ~Qt.ItemIsEditable)
            self._stats_table.setItem(row, 1, group_item)

            for col_offset, (metric_key, _) in enumerate(metrics):
                result = self.statistics_service.calculate_for_workspace_sample(
                    ws,
                    population_name=population_name,
                    channel_name=channel_name,
                    statistic_key=metric_key,
                )
                value_str = self.statistics_service.format_result(result)
                cell = QTableWidgetItem(value_str)
                cell.setBackground(_STAT_BG)
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                cell.setFlags(cell.flags() & ~Qt.ItemIsEditable)
                self._stats_table.setItem(row, 2 + col_offset, cell)

        self._stats_table.resizeColumnsToContents()
        self._stats_table.blockSignals(False)

    def _on_export_stats_csv(self) -> None:
        if self._stats_table.rowCount() == 0:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Statistics CSV", "", "CSV files (*.csv)"
        )
        if not path:
            return

        headers = [
            self._stats_table.horizontalHeaderItem(col).text()
            for col in range(self._stats_table.columnCount())
        ]
        rows = [
            [
                (self._stats_table.item(row, col).text()
                 if self._stats_table.item(row, col) else "")
                for col in range(self._stats_table.columnCount())
            ]
            for row in range(self._stats_table.rowCount())
        ]

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

    def _on_export_stats_xlsx(self) -> None:
        if self._stats_table.rowCount() == 0:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Statistics Excel", "", "Excel files (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        # Collect table data
        n_cols = self._stats_table.columnCount()
        n_rows = self._stats_table.rowCount()
        headers = [
            self._stats_table.horizontalHeaderItem(col).text() for col in range(n_cols)
        ]
        rows = [
            [
                (self._stats_table.item(row, col).text()
                 if self._stats_table.item(row, col) else "")
                for col in range(n_cols)
            ]
            for row in range(n_rows)
        ]

        try:
            _write_stats_xlsx(path, headers, rows)
        except Exception as exc:
            QMessageBox.warning(self, "Export error", str(exc))

    # ------------------------------------------------------------------
    # Events tab logic
    # ------------------------------------------------------------------

    def _on_ev_group_changed(self) -> None:
        self._rebuild_ev_gates()
        self._rebuild_ev_sample_list()

    def _rebuild_ev_gates(self) -> None:
        group_name = self._ev_group_combo.currentData()
        current_gate = self._ev_gate_combo.currentText()

        populations, _ = self._collect_groups_populations_channels(group_name)

        self._ev_gate_combo.blockSignals(True)
        self._ev_gate_combo.clear()
        self._ev_gate_combo.addItems(populations)
        idx = self._ev_gate_combo.findText(current_gate)
        if idx >= 0:
            self._ev_gate_combo.setCurrentIndex(idx)
        self._ev_gate_combo.blockSignals(False)

    def _get_gate_mask(self, ws, gate_name: str) -> np.ndarray | None:
        """Return boolean mask for the given gate/population, or None if not found."""
        if gate_name == "All events":
            return np.ones(ws.sample.event_count, dtype=bool)
        gate = next((g for g in ws.gates if g.name == gate_name), None)
        if gate is None:
            return None
        return gate.full_mask

    def _rebuild_ev_sample_list(self) -> None:
        group_name = self._ev_group_combo.currentData()
        gate_name = self._ev_gate_combo.currentText()

        samples = self._non_compensation_samples(group_name)

        self._ev_sample_table.blockSignals(True)
        self._ev_sample_table.setRowCount(len(samples))

        for row, (_, ws) in enumerate(samples):
            mask = self._get_gate_mask(ws, gate_name) if gate_name else None
            has_gate = mask is not None
            event_count = int(mask.sum()) if has_gate else None

            # Checkbox in sample name cell
            name_item = QTableWidgetItem(ws.sample_name)
            name_item.setFlags(name_item.flags() | Qt.ItemIsUserCheckable)
            name_item.setCheckState(Qt.Checked if has_gate else Qt.Unchecked)
            if not has_gate:
                name_item.setFlags(name_item.flags() & ~Qt.ItemIsEnabled)
                name_item.setForeground(_MISSING_FG)
                name_item.setBackground(_MISSING_BG)
            else:
                name_item.setBackground(_SAMPLE_BG)
            self._ev_sample_table.setItem(row, 0, name_item)

            group_item = QTableWidgetItem(ws.group_name)
            group_item.setFlags(group_item.flags() & ~Qt.ItemIsEditable)
            if not has_gate:
                group_item.setForeground(_MISSING_FG)
                group_item.setBackground(_MISSING_BG)
            else:
                group_item.setBackground(_GROUP_BG)
            self._ev_sample_table.setItem(row, 1, group_item)

            count_text = f"{event_count:,}" if event_count is not None else "—"
            count_item = QTableWidgetItem(count_text)
            count_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            count_item.setFlags(count_item.flags() & ~Qt.ItemIsEditable)
            if not has_gate:
                count_item.setForeground(_MISSING_FG)
                count_item.setBackground(_MISSING_BG)
            else:
                count_item.setBackground(_SAMPLE_BG)
            self._ev_sample_table.setItem(row, 2, count_item)

        self._ev_sample_table.resizeColumnsToContents()
        self._ev_sample_table.blockSignals(False)
        self._update_ev_summary()

    def _checked_ev_rows(self) -> list[int]:
        """Return row indices where the sample checkbox is checked."""
        rows = []
        for row in range(self._ev_sample_table.rowCount()):
            item = self._ev_sample_table.item(row, 0)
            if item and item.checkState() == Qt.Checked:
                rows.append(row)
        return rows

    def _update_ev_summary(self) -> None:
        gate_name = self._ev_gate_combo.currentText() or "—"
        group_name = self._ev_group_combo.currentData()
        checked = self._checked_ev_rows()
        samples = self._non_compensation_samples(group_name)

        total_events = 0
        for row in checked:
            if row >= len(samples):
                continue
            _, ws = samples[row]
            mask = self._get_gate_mask(ws, gate_name)
            if mask is not None:
                total_events += int(mask.sum())

        if checked:
            self._ev_summary_label.setText(
                f"<b>{len(checked)}</b> muestra(s)<br>"
                f"Gate: <b>{gate_name}</b><br>"
                f"Total: <b>{total_events:,}</b> eventos"
            )
        else:
            self._ev_summary_label.setText("Sin muestras seleccionadas")

    def _on_ev_select_all(self) -> None:
        self._ev_sample_table.blockSignals(True)
        for row in range(self._ev_sample_table.rowCount()):
            item = self._ev_sample_table.item(row, 0)
            if item and (item.flags() & Qt.ItemIsEnabled):
                item.setCheckState(Qt.Checked)
        self._ev_sample_table.blockSignals(False)
        self._update_ev_summary()

    def _on_ev_deselect_all(self) -> None:
        self._ev_sample_table.blockSignals(True)
        for row in range(self._ev_sample_table.rowCount()):
            item = self._ev_sample_table.item(row, 0)
            if item:
                item.setCheckState(Qt.Unchecked)
        self._ev_sample_table.blockSignals(False)
        self._update_ev_summary()

    def _on_ev_select_ready(self) -> None:
        """Check only samples that have the selected gate."""
        self._ev_sample_table.blockSignals(True)
        for row in range(self._ev_sample_table.rowCount()):
            item = self._ev_sample_table.item(row, 0)
            if item:
                has_gate = bool(item.flags() & Qt.ItemIsEnabled)
                item.setCheckState(Qt.Checked if has_gate else Qt.Unchecked)
        self._ev_sample_table.blockSignals(False)
        self._update_ev_summary()

    # ── Events export ─────────────────────────────────────────────────

    def _selected_ev_samples_with_masks(
        self,
    ) -> list[tuple[object, np.ndarray]]:
        """Return (WorkspaceSample, mask) for each checked row that has a gate."""
        gate_name = self._ev_gate_combo.currentText()
        group_name = self._ev_group_combo.currentData()
        samples = self._non_compensation_samples(group_name)
        result = []
        for row in self._checked_ev_rows():
            if row >= len(samples):
                continue
            _, ws = samples[row]
            mask = self._get_gate_mask(ws, gate_name)
            if mask is not None:
                result.append((ws, mask))
        return result

    def _on_export_events_csv(self) -> None:
        pairs = self._selected_ev_samples_with_masks()
        if not pairs:
            QMessageBox.information(self, "Export", "No hay muestras seleccionadas para exportar.")
            return

        gate_name = self._ev_gate_combo.currentText() or "events"
        combined = self._ev_radio_combined.isChecked()

        if combined:
            path, _ = QFileDialog.getSaveFileName(
                self, "Export Events CSV", "", "CSV files (*.csv)"
            )
            if not path:
                return

            # Build combined CSV with a leading "sample" column
            all_rows: list[list] = []
            headers: list[str] | None = None
            for ws, mask in pairs:
                selected = ws.sample.effective_events[mask]
                col_names = [ch.display_name for ch in ws.sample.channels]
                if headers is None:
                    headers = ["sample"] + col_names
                for event_row in selected:
                    all_rows.append([ws.sample_name] + list(event_row))

            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers or ["sample"])
                writer.writerows(all_rows)

        else:
            directory = QFileDialog.getExistingDirectory(
                self, "Select output directory"
            )
            if not directory:
                return

            out_dir = Path(directory)
            gate_part = _safe_filename(gate_name)
            errors: list[str] = []
            for ws, mask in pairs:
                sample_part = _safe_filename(ws.sample_name)
                out_path = out_dir / f"{sample_part}_{gate_part}.csv"
                try:
                    export_masked_events_to_csv(ws.sample, mask, out_path)
                except Exception as exc:
                    errors.append(f"{ws.sample_name}: {exc}")

            if errors:
                QMessageBox.warning(
                    self, "Export parcial", "Errores en algunas muestras:\n" + "\n".join(errors)
                )
            else:
                QMessageBox.information(
                    self,
                    "Export completado",
                    f"{len(pairs)} archivo(s) guardados en:\n{directory}",
                )

    def _on_export_events_fcs(self) -> None:
        pairs = self._selected_ev_samples_with_masks()
        if not pairs:
            QMessageBox.information(self, "Export", "No hay muestras seleccionadas para exportar.")
            return

        gate_name = self._ev_gate_combo.currentText() or "events"
        directory = QFileDialog.getExistingDirectory(
            self, "Select output directory for FCS files"
        )
        if not directory:
            return

        out_dir = Path(directory)
        gate_part = _safe_filename(gate_name)
        errors: list[str] = []
        for ws, mask in pairs:
            sample_part = _safe_filename(ws.sample_name)
            out_path = out_dir / f"{sample_part}_{gate_part}.fcs"
            try:
                export_masked_events_to_fcs(ws.sample, mask, out_path)
            except Exception as exc:
                errors.append(f"{ws.sample_name}: {exc}")

        if errors:
            QMessageBox.warning(
                self, "Export parcial", "Errores en algunas muestras:\n" + "\n".join(errors)
            )
        else:
            QMessageBox.information(
                self,
                "Export completado",
                f"{len(pairs)} archivo(s) FCS guardados en:\n{directory}",
            )
